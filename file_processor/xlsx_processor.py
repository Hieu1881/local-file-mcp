import openpyxl
from typing import Dict


class XLSXProcessor:
    def __init__(self, max_sheets: int = 5, max_rows_per_sheet: int = 50):
        self.max_sheets = max_sheets
        self.max_rows_per_sheet = max_rows_per_sheet

    def extract_text_only(self, file_path: str) -> Dict[str, str]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        text_content = {}
        sheets = list(wb.worksheets)
        for i, sheet in enumerate(sheets[: self.max_sheets]):
            rows = []
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_idx > self.max_rows_per_sheet:
                    break
                row_vals = ["" if v is None else str(v) for v in row]
                rows.append("\t".join(row_vals))
            text_content[f"sheet_{i+1}_{sheet.title}"] = "\n".join(rows)
        return {"text_content": text_content}

    def extract_metadata_only(self, file_path: str) -> Dict[str, Dict[str, str]]:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        props = wb.properties
        metadata = {
            "title": props.title,
            "creator": props.creator,
            "created": str(props.created) if props.created else None,
            "modified": str(props.modified) if props.modified else None,
        }
        return {"metadata": metadata}

    def extract_text_and_metadata(self, file_path: str) -> Dict[str, Dict[str, str]]:
        meta = self.extract_metadata_only(file_path)["metadata"]
        text = self.extract_text_only(file_path)["text_content"]
        return {"metadata": meta, "text_content": text}
